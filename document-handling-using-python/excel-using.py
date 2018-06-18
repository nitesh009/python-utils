
import openpyxl
import os

os.chdir('/Users/Username/Downloads/')

workbook = openpyxl.load_workbook('example.xlsx')

print(type(workbook))

sheet = workbook.get_sheet_by_name('Sheet1')
print(sheet)

sheets = workbook.get_sheet_names()
print(sheets)

cell = sheet['A1']
print(cell.value)

print(sheet['B1'].value)

sheet.cell(row=1, column=2)

for i in range(1,8):
    print(i, sheet.cell(row=i,column=2).value)
